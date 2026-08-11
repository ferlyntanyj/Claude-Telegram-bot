"""
Build the final formatted Excel deliverable for the SGX Liquidity Momentum
Screener: Top 10 sheet, full liquid-universe sheet, and a methodology sheet.
"""
import datetime as dt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TOP10_FILL = PatternFill("solid", fgColor="FFF2CC")
BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
SUB_FONT = Font(name=FONT_NAME, italic=True, size=10, color="595959")

liquid = pd.read_csv("../output/liquidity_momentum_screener_liquid.csv")
full = pd.read_csv("../output/liquidity_momentum_screener.csv")
run_date = dt.date.today().isoformat()
latest_trading_date = liquid["recent_end"].max()
recent_start = liquid["recent_start"].min()

wb = Workbook()

COLUMNS = [
    ("yahoo_ticker", "Ticker", 12),
    ("name", "Company", 28),
    ("market", "Board", 11),
    ("surge_ratio", "Surge Ratio (5D ADTV / 2024 ADTV)", 14),
    ("adtv_recent", "Trailing 5D ADTV (SGD)", 18),
    ("adtv_2024", "2024 Avg Daily Traded Value (SGD)", 20),
    ("n_days_2024", "2024 Trading Days Used", 12),
]


def write_sheet(ws, df, title, subtitle, highlight_top10):
    ws.merge_cells("A1:G1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A2:G2")
    ws["A2"] = subtitle
    ws["A2"].font = SUB_FONT

    header_row = 4
    for i, (_, label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=i, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[header_row].height = 32

    for r, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        ws.cell(row=r, column=1, value=row["yahoo_ticker"]).font = Font(name=FONT_NAME)
        ws.cell(row=r, column=2, value=row["name"]).font = Font(name=FONT_NAME)
        ws.cell(row=r, column=3, value=row["market"]).font = Font(name=FONT_NAME)

        recent_cell = ws.cell(row=r, column=5, value=round(float(row["adtv_recent"]), 2))
        recent_cell.number_format = "$#,##0"
        recent_cell.font = Font(name=FONT_NAME)

        base_cell = ws.cell(row=r, column=6, value=round(float(row["adtv_2024"]), 2))
        base_cell.number_format = "$#,##0"
        base_cell.font = Font(name=FONT_NAME)

        ratio_cell = ws.cell(row=r, column=4, value=f"=E{r}/F{r}")
        ratio_cell.number_format = '0.00"x"'
        ratio_cell.font = Font(name=FONT_NAME, bold=True)

        days_cell = ws.cell(row=r, column=7, value=int(row["n_days_2024"]))
        days_cell.font = Font(name=FONT_NAME)
        days_cell.alignment = Alignment(horizontal="center")

        for c in range(1, 8):
            ws.cell(row=r, column=c).border = BORDER

        if highlight_top10 and r - header_row <= 10:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = TOP10_FILL

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:G{header_row + len(df)}"


# --- Sheet 1: Top 10 ---
ws1 = wb.active
ws1.title = "Top 10"
top10 = liquid.head(10)
write_sheet(
    ws1, top10,
    "SGX Liquidity Momentum Screener — Top 10",
    f"Ranked by ADTV surge ratio (trailing 5-trading-day ADTV ÷ full-year 2024 ADTV). "
    f"Recent window: {recent_start} to {latest_trading_date}. Liquidity floor: 2024 ADTV ≥ SGD 100,000/day. Run date: {run_date}.",
    highlight_top10=True,
)

# --- Sheet 2: Liquid universe (110 names) ---
ws2 = wb.create_sheet("Liquid Universe")
write_sheet(
    ws2, liquid,
    "SGX Liquidity Momentum Screener — Liquid Universe",
    f"All {len(liquid)} SGX mainboard/Catalist stocks with 2024 ADTV ≥ SGD 100,000/day, ranked by surge ratio. "
    f"Recent window: {recent_start} to {latest_trading_date}.",
    highlight_top10=True,
)

# --- Sheet 3: Full universe (unfiltered, 495 names) ---
ws3 = wb.create_sheet("Full Universe (Unfiltered)")
write_sheet(
    ws3, full,
    "SGX Liquidity Momentum Screener — Full Universe (No Liquidity Floor)",
    f"All {len(full)} SGX mainboard/Catalist stocks with sufficient 2024 history, ranked by surge ratio. "
    f"CAUTION: names with very low 2024 ADTV can show extreme ratios from a single small trade, not genuine momentum. "
    f"Recent window: {recent_start} to {latest_trading_date}.",
    highlight_top10=False,
)

# --- Sheet 4: Methodology ---
ws4 = wb.create_sheet("Methodology")
ws4.column_dimensions["A"].width = 100
notes = [
    ("SGX Liquidity Momentum Screener — Methodology", TITLE_FONT),
    ("", None),
    ("Universe", Font(name=FONT_NAME, bold=True)),
    ("All SGX Mainboard and Catalist-listed common stocks (excludes REITs, business trusts, ETFs, "
     "warrants, DLCs, bonds). Source: SGX public securities API (api.sgx.com/securities/v1.1), pulled "
     f"on {run_date}. 565 tickers identified; 558 had retrievable Yahoo Finance price history.", None),
    ("", None),
    ("Data", Font(name=FONT_NAME, bold=True)),
    ("Daily close price and volume from Yahoo Finance (tickers suffixed .SI), 2024-01-01 to present. "
     "Traded value for a given day = close price × volume (unadjusted, so it reflects actual SGD "
     "value traded that day).", None),
    ("", None),
    ("Formula", Font(name=FONT_NAME, bold=True)),
    ("ADTV 2024 = mean daily traded value across all 2024 trading days (requires ≥ 100 trading days "
     "of 2024 data to be included, to avoid partial-year distortion).", None),
    (f"ADTV Recent = mean daily traded value over the trailing 5 trading days ({recent_start} to "
     f"{latest_trading_date}).", None),
    ("Surge Ratio = ADTV Recent ÷ ADTV 2024. A ratio of 5.0x means the stock is trading roughly 5 "
     "times its average 2024 daily value over the last week.", None),
    ("", None),
    ("Liquidity floor", Font(name=FONT_NAME, bold=True)),
    ("Stocks with 2024 ADTV below SGD 100,000/day are excluded from the 'Top 10' and 'Liquid Universe' "
     "sheets. Without this floor, near-untraded microcaps dominate the top of the ranking purely because "
     "a tiny historical base (e.g. SGD 50/day) makes any single trade register as a 100x+ 'surge' — "
     "this is noise, not liquidity momentum. The unfiltered ranking is kept on the 'Full Universe' sheet "
     "for reference. The SGD 100,000 threshold is a judgment call, not a market convention — adjust "
     "in scripts/03_compute_screener.py (MIN_ADTV_2024_SGD) if a different bar is wanted.", None),
    ("", None),
    ("Caveats", Font(name=FONT_NAME, bold=True)),
    ("Yahoo Finance data for SGX counters can have gaps or reporting lag for thinly-traded names. "
     "Corporate actions (splits, bonus issues) do not distort traded-value figures since price and "
     "volume move inversely, but a stock that recently listed, delisted, or was suspended for part of "
     "2024 may show a skewed base ADTV. Ratios are a screen for further research, not a standalone "
     "trading signal.", None),
    ("", None),
    ("Reproduce", Font(name=FONT_NAME, bold=True)),
    ("Run scripts/01_get_universe.py → 02_fetch_history.py → 03_compute_screener.py → "
     "04_build_workbook.py in order (Python 3.12, packages: pandas, yfinance, openpyxl, pyarrow, requests).", None),
]
r = 1
for text, font in notes:
    cell = ws4.cell(row=r, column=1, value=text)
    cell.font = font if font else Font(name=FONT_NAME)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws4.row_dimensions[r].height = 30 if text and len(text) > 80 else 18
    r += 1

wb.save("../output/SGX_Liquidity_Momentum_Screener.xlsx")
print("Saved ../output/SGX_Liquidity_Momentum_Screener.xlsx")
